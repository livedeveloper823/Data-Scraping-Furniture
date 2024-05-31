from time import sleep
from selenium import webdriver
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.common.by import By
import openpyxl, json, re, os

def wait_url(driver:webdriver.Chrome, url: str):
    while True:
        cur_url = driver.current_url
        if cur_url == url:
            break
        sleep(0.1)

def find_element(driver: webdriver.Chrome, whichBy, unique: str) -> WebElement:
    while True:
        try:
            element = driver.find_element(whichBy, unique)
            break
        except:
            pass
        sleep(1)
    return element

def find_elements(driver : webdriver.Chrome, whichBy, unique: str) -> list[WebElement]:
    while True:
        try:
            elements =driver.find_elements(whichBy, unique)
            break
        except:
            pass
        sleep(1)
    return elements

wb = openpyxl.Workbook()
ws = wb.active

ws['A1'] = "Categoría"
ws['B1'] = "Línea"
ws['C1'] = "Sublinea"
ws['D1'] = "Atributos1"
ws['E1'] = "Atributos2"
ws['F1'] = "Atributos3"
ws['G1'] = "Atributos4"
ws['H1'] = "Atributos5"
ws['I1'] = "Atributos6"
ws['J1'] = "Atributos7"
ws['K1'] = "Atributos8"
ws['L1'] = "Atributos9"
ws['M1'] = "Atributos10"
ws['N1'] = "Atributos11"
ws['O1'] = "Atributos11"
ws['P1'] = "Atributos11"
ws['Q1'] = "Atributos11"
ws['R1'] = "Comparables para Modelo"


wb.save("aristas.xlsx")

match_num = 0
category = ""
products_name = ""
price1 = ""
price2 = ""
product_size = ""
product_color = ""
product_fabric = ""
product_type = ""
product_origin = ""
product_certification = ""
product_structure = ""
delivery_time = ""
prodcut_caution = ""
img_url = ""

whole_data =[]

urls = [
    "https://aristas.co/categoria-producto/salas/",
    "https://aristas.co/categoria-producto/comedores/",
    "https://aristas.co/categoria-producto/dormitorios/",
    "https://aristas.co/categoria-producto/estudios/",
    "https://aristas.co/categoria-producto/accesorios/",
    ]


driver = webdriver.Chrome()
driver.maximize_window()
for url in urls:
    driver.get(url)
    wait_url(driver, url)

    categories =[]
    try:
        sub_categories = driver.find_elements(By.CLASS_NAME, "subcategory-divider")
        for sub_category in sub_categories:
            category = sub_category.text
            categories.append(category)
    except:
        pass

    try :
        products_groups = driver.find_elements(By.CLASS_NAME, "products")
    except:
        pass
    category_num = 0
    if not sub_categories:
        category_num = 1
    else: category_num = len(sub_categories)
    print(category_num)
    workbook = openpyxl.load_workbook("aristas.xlsx")
    sheet = workbook['Sheet']
    for i in range(category_num):
        product_urls =[]
        product_prices = []
        products_groups = find_elements(driver, By.CLASS_NAME, "products")
        category_products = products_groups[i].find_elements(By.CLASS_NAME, "product-type-variable")
        for category_product in category_products:
            category_product_url = category_product.find_element(By.CLASS_NAME, "image-fade_in_back").find_element(By.TAG_NAME, "a")
            product_link = category_product_url.get_attribute("href")
            product_urls.append(product_link)
            product_price = category_product.find_element(By.CLASS_NAME, "price-wrapper").text
            product_prices.append(product_price)
        for j in range(len(product_urls)):
            match_num += 1
            driver.get(product_urls[j])
            print(product_prices[j].replace("$","").replace(",", "").replace(" ", ""))
            if "–" in product_prices[j].replace("$","").replace(",", "").replace(" ", ""):
                range_price = re.findall(r'\d+', product_prices[j].replace("$","").replace(",", "").replace(" ", ""))
                print(range_price)
                sheet[f'F{match_num + 1}'] = range_price[0]
                sheet[f'G{match_num + 1}'] = range_price[1]
                # for json
                price1 = range_price[0]
                price2 = range_price[1]
            else:
                length = len(product_prices[j].replace("$","").replace(",", "").replace(" ", ""))
                fixed_price = product_prices[j].replace("$","").replace(",", "").replace(" ", "")[int(length/2):]
                sheet[f'G{match_num + 1}'] = fixed_price
                price2 = range_price[1]
            print(product_urls[j])
            print(product_prices[j].replace("$","").replace(",", "").replace(" ", ""))
            sheet[f'C{match_num + 1}'] = categories[i]
            category = categories[i]
            product_name = find_element(driver, By.CLASS_NAME, "product_title").text
            sheet[f'D{match_num + 1}'] = product_name
            products_name = product_name
            info_table = find_element(driver, By.CSS_SELECTOR, "#tab-additional_information > table")
            sleep(0.5)
            info_fields = info_table.find_elements(By.XPATH, ".//tr")
            print(len(info_fields))
            for info_field in info_fields:
                product_title = info_field.find_element(By.XPATH, ".//th").text
                print(product_title)
                if product_title == "TIEMPO DE ENTREGA":
                    sheet[f'H{match_num + 1}'] = info_field.find_element(By.XPATH, ".//td").text
                    delivery_time = info_field.find_element(By.XPATH, ".//td").text
                elif product_title == "MEDIDAS":
                    sheet[f'I{match_num + 1}'] = info_field.find_element(By.XPATH, ".//td").text
                    product_size = info_field.find_element(By.XPATH, ".//td").text
                elif product_title == "MADERA Y PINTURA":
                    sheet[f'J{match_num + 1}'] = info_field.find_element(By.XPATH, ".//td").text
                    product_color = info_field.find_element(By.XPATH, ".//td").text
                elif product_title == "ESPUMA":
                    sheet[f'K{match_num + 1}'] = info_field.find_element(By.XPATH, ".//td").text
                    product_type = info_field.find_element(By.XPATH, ".//td").text
                elif product_title == "TELA":
                    sheet[f'L{match_num + 1}'] = info_field.find_element(By.XPATH, ".//td").text
                    product_fabric = info_field.find_element(By.XPATH, ".//td").text
                elif product_title == "ESTRUCTURE":
                    sheet[f'M{match_num + 1}'] = info_field.find_element(By.XPATH, ".//td").text
                    product_structure = info_field.find_element(By.XPATH, ".//td").text
                elif product_title == "CERTIFICACIONES":
                    sheet[f'N{match_num + 1}'] = info_field.find_element(By.XPATH, ".//td").text
                    product_certification = info_field.find_element(By.XPATH, ".//td").text
                elif product_title == "CUIDADO":
                    sheet[f'O{match_num + 1}'] = info_field.find_element(By.XPATH, ".//td").text
                    prodcut_caution = info_field.find_element(By.XPATH, ".//td").text
                elif product_title == "ORIGEN":
                    sheet[f'P{match_num + 1}'] = info_field.find_element(By.XPATH, ".//td").text
                    product_origin = info_field.find_element(By.XPATH, ".//td").text
            product_img = find_element(driver, By.CLASS_NAME, "jet-woo-product-gallery__image")
            image_url = product_img.find_element(By.TAG_NAME, "img").get_attribute("src")
            sheet[f'Q{match_num + 1}'] = image_url
            img_url = image_url
            workbook.save("aristas.xlsx")
            print(category)
            data = {
                "category": category,
                "product information": {
                "product_name":product_name,
                "price1":price1,
                "price2":price2,
                "delivery_time":delivery_time,
                "product_size":product_size,
                "product_color":product_color,
                "product_type":product_type,
                "product_fabric":product_fabric,
                "product_structure":product_structure,
                "product_certification":product_certification,
                "product_caution":prodcut_caution,
                "product_origin":product_origin,
                "image_url":img_url
                }
            }
            whole_data.append(data)
        driver.get(url)
        match_num +=2
        sleep(1)
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        sleep(3)
        

with open("aristas.json", 'w') as file:
    json.dump(whole_data, file)

